import sys
#!/usr/bin/env python3
"""KairoLogic Safe Harbor Evidence Ledger - Professional XLSX"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.comments import Comment
import subprocess

NAVY = '0B1E3D'
NAVY_LIGHT = '1A3A5F'
GOLD = 'D4A574'
GOLD_DARK = 'B88F5F'
WHITE = 'FFFFFF'
GRAY_50 = 'F9FAFB'
GRAY_100 = 'F3F4F6'
GRAY_200 = 'E5E7EB'
GRAY_500 = '6B7280'
GRAY_700 = '374151'
RED_50 = 'FEF2F2'
RED_600 = 'DC2626'
GREEN_50 = 'F0FDF4'
GREEN_600 = '059669'
AMBER_50 = 'FFFBEB'
AMBER_600 = 'D97706'
BLUE_50 = 'EFF6FF'

thin_border = Border(
    left=Side(style='thin', color=GRAY_200),
    right=Side(style='thin', color=GRAY_200),
    top=Side(style='thin', color=GRAY_200),
    bottom=Side(style='thin', color=GRAY_200))

header_border = Border(
    left=Side(style='thin', color=NAVY_LIGHT),
    right=Side(style='thin', color=NAVY_LIGHT),
    top=Side(style='thin', color=NAVY_LIGHT),
    bottom=Side(style='medium', color=GOLD_DARK))

header_font = Font(name='Arial', size=10, bold=True, color=WHITE)
header_fill = PatternFill('solid', fgColor=NAVY)
header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

body_font = Font(name='Arial', size=10, color=GRAY_700)
body_align = Alignment(vertical='center', wrap_text=True)

title_font = Font(name='Arial', size=14, bold=True, color=NAVY)
subtitle_font = Font(name='Arial', size=10, color=GRAY_500)
section_font = Font(name='Arial', size=11, bold=True, color=NAVY_LIGHT)

gold_fill = PatternFill('solid', fgColor=GOLD)
alt_fill = PatternFill('solid', fgColor=GRAY_50)
white_fill = PatternFill('solid', fgColor=WHITE)
green_fill = PatternFill('solid', fgColor=GREEN_50)
red_fill = PatternFill('solid', fgColor=RED_50)
amber_fill = PatternFill('solid', fgColor=AMBER_50)
blue_fill = PatternFill('solid', fgColor=BLUE_50)

example_font = Font(name='Arial', size=10, italic=True, color=GRAY_500)


def style_header_row(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = header_border


def style_data_row(ws, row, cols, is_example=False, is_alt=False):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = example_font if is_example else body_font
        cell.fill = alt_fill if is_alt else white_fill
        cell.alignment = body_align
        cell.border = thin_border


def add_title_block(ws, title, subtitle, start_row=1):
    ws.cell(row=start_row, column=1, value=title).font = title_font
    ws.cell(row=start_row + 1, column=1, value=subtitle).font = subtitle_font
    # Gold accent bar
    for c in range(1, 12):
        ws.cell(row=start_row + 2, column=c).fill = PatternFill('solid', fgColor=GOLD)
        ws.cell(row=start_row + 2, column=c).border = Border(bottom=Side(style='thin', color=GOLD_DARK))
    ws.row_dimensions[start_row + 2].height = 3
    return start_row + 3


def build_tab1_vendor_inventory(wb):
    ws = wb.active
    ws.title = 'Digital Supply Chain'
    ws.sheet_properties.tabColor = NAVY

    r = add_title_block(ws,
        'DIGITAL SUPPLY CHAIN INVENTORY',
        'SB 1188 Vendor Due Diligence Register  |  [Practice Name]  |  Effective: February 2026')

    # Instructions
    r += 1
    ws.cell(row=r, column=1, value='INSTRUCTIONS: Document every vendor that accesses, stores, processes, or transmits patient data (PHI/PII). Update quarterly. Rows in italics are examples \u2014 replace with your actual vendors.').font = Font(name='Arial', size=9, italic=True, color=GRAY_500)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=11)
    r += 2

    headers = [
        '#', 'Vendor Name', 'Service Provided', 'Risk Tier',
        'PHI Access?', 'Server Location', 'Cloud Region',
        'Verification Date', 'Proof Type', 'Verified By', 'Status'
    ]
    widths = [5, 22, 24, 14, 12, 20, 16, 16, 18, 16, 16]

    header_row = r
    for c, (h, w) in enumerate(zip(headers, widths), 1):
        ws.cell(row=r, column=c, value=h)
        ws.column_dimensions[get_column_letter(c)].width = w
    style_header_row(ws, r, len(headers))
    r += 1

    # Example data
    examples = [
        ['1', 'eClinicalWorks', 'EMR / Patient Portal', 'CRITICAL', 'Yes', 'USA (Texas / Virginia)', 'AWS us-east-1', '01/15/2026', 'SOC 2 Type II Report', 'J. Martinez', 'SOVEREIGN'],
        ['2', 'Google Workspace', 'Email / Calendar / Drive', 'HIGH', 'Yes', 'USA (Multiple)', 'GCP us-central1', '01/20/2026', 'DPA + Data Residency Cert', 'J. Martinez', 'SOVEREIGN'],
        ['3', 'Mailchimp (Intuit)', 'Patient Newsletter', 'MODERATE', 'No (PII only)', 'USA (Atlanta, GA)', 'AWS us-east-1', '02/01/2026', 'TOS Review + Email Conf.', 'S. Lee', 'SOVEREIGN'],
        ['4', 'Freed Health', 'AI Clinical Transcription', 'CRITICAL', 'Yes', 'USA (Verified)', 'AWS us-east-2', '02/05/2026', 'Vendor Certificate + NPI Audit', 'J. Martinez', 'SOVEREIGN'],
        ['5', 'Doxy.me', 'Telehealth Platform', 'CRITICAL', 'Yes', 'USA (Virginia)', 'AWS us-east-1', '02/03/2026', 'BAA + Server Location Letter', 'S. Lee', 'SOVEREIGN'],
        ['6', 'WP Engine', 'Website Hosting', 'HIGH', 'No (PII only)', 'USA (Texas)', 'GCP us-central1', '01/25/2026', 'DNS + IP Geolocation Audit', 'IT Admin', 'SOVEREIGN'],
        ['7', 'Weave Communications', 'Patient Messaging / VoIP', 'HIGH', 'Yes', 'USA (Utah)', 'AWS us-west-2', '02/01/2026', 'Vendor Email Confirmation', 'S. Lee', 'SOVEREIGN'],
        ['8', 'Waystar', 'Billing / Revenue Cycle', 'HIGH', 'Yes', 'USA (Kentucky)', 'Private Cloud (US)', '01/28/2026', 'SOC 2 Report + BAA', 'J. Martinez', 'SOVEREIGN'],
        ['9', '[New Vendor]', '[Service]', '[Tier]', '[Yes/No]', '[Location]', '[Region]', '', '', '', 'PENDING'],
        ['10', '', '', '', '', '', '', '', '', '', ''],
    ]

    for i, row_data in enumerate(examples):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)
        is_example = i < 8
        style_data_row(ws, r, len(headers), is_example=(i < 8), is_alt=(i % 2 == 1))

        # Color the Status column
        status_cell = ws.cell(row=r, column=11)
        if 'SOVEREIGN' in str(status_cell.value):
            status_cell.fill = green_fill
            status_cell.font = Font(name='Arial', size=10, bold=True, color=GREEN_600)
        elif 'PENDING' in str(status_cell.value or ''):
            status_cell.fill = amber_fill
            status_cell.font = Font(name='Arial', size=10, bold=True, color=AMBER_600)

        # Color Risk Tier
        tier_cell = ws.cell(row=r, column=4)
        tier_val = str(tier_cell.value or '')
        if 'CRITICAL' in tier_val:
            tier_cell.font = Font(name='Arial', size=10, bold=True, color=RED_600)
        elif 'HIGH' in tier_val:
            tier_cell.font = Font(name='Arial', size=10, bold=True, color=AMBER_600)

        r += 1

    # Add 20 more blank rows
    for i in range(20):
        row_num = r + i
        ws.cell(row=row_num, column=1, value=str(11 + i))
        for c in range(1, len(headers) + 1):
            style_data_row(ws, row_num, len(headers), is_alt=(i % 2 == 1))

    # Data validations
    dv_phi = DataValidation(type='list', formula1='"Yes,No (PII only),No,Pending Review"', allow_blank=True)
    dv_phi.error = 'Select from list'
    dv_phi.prompt = 'Does this vendor access PHI?'
    ws.add_data_validation(dv_phi)
    dv_phi.add(f'E{header_row+1}:E{r+20}')

    dv_tier = DataValidation(type='list', formula1='"CRITICAL,HIGH,MODERATE,LOW"', allow_blank=True)
    ws.add_data_validation(dv_tier)
    dv_tier.add(f'D{header_row+1}:D{r+20}')

    dv_status = DataValidation(type='list', formula1='"SOVEREIGN,PENDING,UNVERIFIED,NON-COMPLIANT,UNDER REVIEW"', allow_blank=True)
    ws.add_data_validation(dv_status)
    dv_status.add(f'K{header_row+1}:K{r+20}')

    dv_proof = DataValidation(type='list', formula1='"SOC 2 Type II Report,SOC 2 Type I Report,Vendor Certificate,BAA + Server Letter,DPA + Residency Cert,DNS/IP Geolocation Audit,Email Confirmation,TOS Review,NPI Audit,Contract Review,Pending"', allow_blank=True)
    ws.add_data_validation(dv_proof)
    dv_proof.add(f'I{header_row+1}:I{r+20}')

    # Summary section
    sr = r + 22
    ws.cell(row=sr, column=1, value='LEDGER SUMMARY').font = section_font
    ws.cell(row=sr + 1, column=1, value='Total Vendors Inventoried:').font = Font(name='Arial', size=10, color=GRAY_700)
    ws.cell(row=sr + 1, column=3, value=f'=COUNTA(B{header_row+1}:B{r+19})-COUNTBLANK(B{header_row+1}:B{r+19})').font = Font(name='Arial', size=10, bold=True, color=NAVY)
    ws.cell(row=sr + 2, column=1, value='Sovereign (Verified):').font = Font(name='Arial', size=10, color=GREEN_600)
    ws.cell(row=sr + 2, column=3, value=f'=COUNTIF(K{header_row+1}:K{r+19},"SOVEREIGN")').font = Font(name='Arial', size=10, bold=True, color=GREEN_600)
    ws.cell(row=sr + 3, column=1, value='Pending Verification:').font = Font(name='Arial', size=10, color=AMBER_600)
    ws.cell(row=sr + 3, column=3, value=f'=COUNTIF(K{header_row+1}:K{r+19},"PENDING")').font = Font(name='Arial', size=10, bold=True, color=AMBER_600)
    ws.cell(row=sr + 4, column=1, value='Non-Compliant / Unverified:').font = Font(name='Arial', size=10, color=RED_600)
    ws.cell(row=sr + 4, column=3, value=f'=COUNTIF(K{header_row+1}:K{r+19},"NON-COMPLIANT")+COUNTIF(K{header_row+1}:K{r+19},"UNVERIFIED")').font = Font(name='Arial', size=10, bold=True, color=RED_600)
    ws.cell(row=sr + 6, column=1, value='Last Updated:').font = Font(name='Arial', size=10, color=GRAY_500)
    ws.cell(row=sr + 6, column=3, value='[Date]').font = Font(name='Arial', size=10, color=NAVY)
    ws.cell(row=sr + 7, column=1, value='Updated By:').font = Font(name='Arial', size=10, color=GRAY_500)
    ws.cell(row=sr + 7, column=3, value='[Name / Title]').font = Font(name='Arial', size=10, color=NAVY)

    # Freeze panes
    ws.freeze_panes = f'A{header_row + 1}'
    ws.auto_filter.ref = f'A{header_row}:K{r+19}'


def build_tab2_residency_signals(wb):
    ws = wb.create_sheet('Technical Residency Signals')
    ws.sheet_properties.tabColor = NAVY_LIGHT

    r = add_title_block(ws,
        'TECHNICAL RESIDENCY SIGNALS',
        'IP Whitelist & Infrastructure Verification Register  |  [Practice Name]')

    r += 1
    ws.cell(row=r, column=1, value='INSTRUCTIONS: Track all network endpoints, devices, and remote access points. Verify that no data routes through foreign proxies, VPNs, or offshore IT support. Update quarterly or upon any infrastructure change.').font = Font(name='Arial', size=9, italic=True, color=GRAY_500)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=10)
    r += 2

    headers = [
        '#', 'Device / System', 'Primary IP Address', 'Authorized Region',
        'IT Provider', 'Connection Type', 'Last Verified', 'Verified By',
        'Foreign Routing Detected?', 'Status'
    ]
    widths = [5, 24, 20, 18, 20, 16, 16, 16, 22, 16]

    header_row = r
    for c, (h, w) in enumerate(zip(headers, widths), 1):
        ws.cell(row=r, column=c, value=h)
        ws.column_dimensions[get_column_letter(c)].width = w
    style_header_row(ws, r, len(headers))
    r += 1

    examples = [
        ['1', 'Main Office Router', '192.168.x.x (Static)', 'Austin, TX', 'Spectrum Business', 'Fiber / Static IP', '02/01/2026', 'IT Admin', 'No', 'SOVEREIGN'],
        ['2', 'Billing Dept VPN', '45.x.x.x (Masked)', 'US-Central', 'Internal IT', 'Site-to-Site VPN', '02/01/2026', 'IT Admin', 'No', 'SOVEREIGN'],
        ['3', 'Remote Admin Access', '[Masked]', 'Dallas, TX', 'SecureOps MSP', 'RDP over VPN', '01/28/2026', 'J. Martinez', 'No', 'SOVEREIGN'],
        ['4', 'Telehealth Endpoint', '[Provider IP]', 'US-East', 'Doxy.me', 'WebRTC (Browser)', '02/03/2026', 'S. Lee', 'No', 'SOVEREIGN'],
        ['5', 'Patient Portal CDN', '[CDN Edge IP]', 'US-Multi', 'Cloudflare (US)', 'HTTPS / CDN', '01/25/2026', 'IT Admin', 'No', 'SOVEREIGN'],
        ['6', 'Staff Wi-Fi Network', '10.x.x.x', 'Austin, TX', 'Internal', 'WPA3 Enterprise', '02/01/2026', 'IT Admin', 'No', 'SOVEREIGN'],
        ['7', 'Backup Replication', '[Backup IP]', 'US-East', 'Datto', 'Encrypted Tunnel', '01/30/2026', 'IT Admin', 'No', 'SOVEREIGN'],
        ['8', '', '', '', '', '', '', '', '', ''],
    ]

    for i, row_data in enumerate(examples):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)
        style_data_row(ws, r, len(headers), is_example=(i < 7), is_alt=(i % 2 == 1))

        status_cell = ws.cell(row=r, column=10)
        if 'SOVEREIGN' in str(status_cell.value or ''):
            status_cell.fill = green_fill
            status_cell.font = Font(name='Arial', size=10, bold=True, color=GREEN_600)

        foreign_cell = ws.cell(row=r, column=9)
        if str(foreign_cell.value or '') == 'No':
            foreign_cell.fill = green_fill

        r += 1

    # 15 more blank rows
    for i in range(15):
        ws.cell(row=r + i, column=1, value=str(9 + i))
        for c in range(1, len(headers) + 1):
            style_data_row(ws, r + i, len(headers), is_alt=(i % 2 == 1))

    # Validations
    dv_foreign = DataValidation(type='list', formula1='"No,Yes - Remediated,Yes - Under Investigation,Unknown"', allow_blank=True)
    ws.add_data_validation(dv_foreign)
    dv_foreign.add(f'I{header_row+1}:I{r+15}')

    dv_status = DataValidation(type='list', formula1='"SOVEREIGN,PENDING,FLAGGED,NON-COMPLIANT"', allow_blank=True)
    ws.add_data_validation(dv_status)
    dv_status.add(f'J{header_row+1}:J{r+15}')

    ws.freeze_panes = f'A{header_row + 1}'
    ws.auto_filter.ref = f'A{header_row}:J{r+14}'


def build_tab3_regulatory_log(wb):
    ws = wb.create_sheet('Regulatory & Cure Notice Log')
    ws.sheet_properties.tabColor = GOLD_DARK

    r = add_title_block(ws,
        'REGULATORY & CURE NOTICE LOG',
        'Active Defense Register  |  [Practice Name]  |  SB 1188 / HB 149 Compliance')

    r += 1
    ws.cell(row=r, column=1, value='INSTRUCTIONS: Document ALL regulatory inquiries, audit requests, patient complaints related to data sovereignty, and Cure Notices received. This log is your primary evidence of "Active Defense" and Reasonable Care. If it isn\'t in the ledger, it didn\'t happen.').font = Font(name='Arial', size=9, italic=True, color=GRAY_500)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=10)
    r += 2

    headers = [
        '#', 'Date Received', 'Agency / Entity', 'Contact Person',
        'Issue Type', 'Issue Description', 'Resolution Action',
        'Evidence Attached?', 'Date Resolved', 'Status'
    ]
    widths = [5, 16, 22, 18, 18, 30, 30, 16, 16, 16]

    header_row = r
    for c, (h, w) in enumerate(zip(headers, widths), 1):
        ws.cell(row=r, column=c, value=h)
        ws.column_dimensions[get_column_letter(c)].width = w
    style_header_row(ws, r, len(headers))
    r += 1

    examples = [
        ['1', '02/08/2026', 'Texas DSHS', 'Inspector R. Gomez', 'AI Transparency Inquiry', 'Routine inquiry regarding AI chatbot on practice website; requested HB 149 disclosure documentation', 'Provided HB 149 AI Disclosure Kit, website screenshot showing footer notice, signed AI consent forms', 'Yes (3 docs)', '02/10/2026', 'RESOLVED'],
        ['2', '02/12/2026', 'Patient Complaint', 'N/A (Anonymous)', 'Data Residency Concern', 'Patient inquired whether their data was stored overseas after seeing AI chatbot', 'Provided patient with AI Transparency Notice; documented conversation in chart; no violation found', 'Yes (1 doc)', '02/12/2026', 'RESOLVED'],
        ['3', '', '', '', '', '', '', '', '', ''],
    ]

    for i, row_data in enumerate(examples):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)
        style_data_row(ws, r, len(headers), is_example=(i < 2), is_alt=(i % 2 == 1))

        status_cell = ws.cell(row=r, column=10)
        if 'RESOLVED' in str(status_cell.value or ''):
            status_cell.fill = green_fill
            status_cell.font = Font(name='Arial', size=10, bold=True, color=GREEN_600)
        elif 'OPEN' in str(status_cell.value or ''):
            status_cell.fill = red_fill
            status_cell.font = Font(name='Arial', size=10, bold=True, color=RED_600)

        r += 1

    for i in range(20):
        ws.cell(row=r + i, column=1, value=str(4 + i))
        for c in range(1, len(headers) + 1):
            style_data_row(ws, r + i, len(headers), is_alt=(i % 2 == 1))

    dv_type = DataValidation(type='list', formula1='"AI Transparency Inquiry,Data Residency Inquiry,Cure Notice,State Audit,Patient Complaint,Vendor Non-Compliance,HIPAA Breach Report,Internal Incident,Other"', allow_blank=True)
    ws.add_data_validation(dv_type)
    dv_type.add(f'E{header_row+1}:E{r+20}')

    dv_status = DataValidation(type='list', formula1='"RESOLVED,OPEN - In Progress,OPEN - Awaiting Response,ESCALATED,PENDING CURE"', allow_blank=True)
    ws.add_data_validation(dv_status)
    dv_status.add(f'J{header_row+1}:J{r+20}')

    dv_evidence = DataValidation(type='list', formula1='"Yes (attached),Yes (on file),Pending,No,N/A"', allow_blank=True)
    ws.add_data_validation(dv_evidence)
    dv_evidence.add(f'H{header_row+1}:H{r+20}')

    ws.freeze_panes = f'A{header_row + 1}'
    ws.auto_filter.ref = f'A{header_row}:J{r+19}'


def build_tab4_quarterly_audit(wb):
    ws = wb.create_sheet('Quarterly Audit Trail')
    ws.sheet_properties.tabColor = GREEN_600

    r = add_title_block(ws,
        'QUARTERLY AUDIT TRAIL',
        'Rolling Compliance Verification Record  |  [Practice Name]')

    r += 1
    ws.cell(row=r, column=1, value='INSTRUCTIONS: Complete one row per quarter. This trail demonstrates ongoing compliance diligence \u2014 the single strongest element of Safe Harbor defense. "If it isn\'t in the ledger, it didn\'t happen."').font = Font(name='Arial', size=9, italic=True, color=GRAY_500)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=9)
    r += 2

    headers = [
        'Quarter', 'Audit Date', 'Sentry Scan\nCompleted?',
        'Vendor Certs\nAll Current?', 'Employee Acks\nAll Current?',
        'Website Disclosure\nVerified?', 'Issues Found', 'Remediation Notes', 'Auditor'
    ]
    widths = [14, 14, 16, 16, 16, 18, 24, 30, 16]

    header_row = r
    for c, (h, w) in enumerate(zip(headers, widths), 1):
        ws.cell(row=r, column=c, value=h)
        ws.column_dimensions[get_column_letter(c)].width = w
    style_header_row(ws, r, len(headers))
    r += 1

    quarters = [
        ['Q1 2026 (Jan-Mar)', '03/31/2026', 'Yes', 'Yes', 'Yes', 'Yes', 'None', 'Initial implementation complete. All vendors verified.', '[Name]'],
        ['Q2 2026 (Apr-Jun)', '', '', '', '', '', '', '', ''],
        ['Q3 2026 (Jul-Sep)', '', '', '', '', '', '', '', ''],
        ['Q4 2026 (Oct-Dec)', '', '', '', '', '', '', '', ''],
        ['Q1 2027 (Jan-Mar)', '', '', '', '', '', '', '', ''],
        ['Q2 2027 (Apr-Jun)', '', '', '', '', '', '', '', ''],
        ['Q3 2027 (Jul-Sep)', '', '', '', '', '', '', '', ''],
        ['Q4 2027 (Oct-Dec)', '', '', '', '', '', '', '', ''],
    ]

    for i, row_data in enumerate(quarters):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)
        style_data_row(ws, r, len(headers), is_example=(i == 0), is_alt=(i % 2 == 1))
        # Highlight "Yes" cells
        for c in [3, 4, 5, 6]:
            cell = ws.cell(row=r, column=c)
            if str(cell.value or '') == 'Yes':
                cell.fill = green_fill
                cell.font = Font(name='Arial', size=10, bold=True, color=GREEN_600)
        r += 1

    dv_yn = DataValidation(type='list', formula1='"Yes,No,Partial,N/A"', allow_blank=True)
    ws.add_data_validation(dv_yn)
    for col in ['C', 'D', 'E', 'F']:
        dv_yn.add(f'{col}{header_row+1}:{col}{r}')

    ws.freeze_panes = f'A{header_row + 1}'


def build_tab5_instructions(wb):
    ws = wb.create_sheet('Instructions & Reference')
    ws.sheet_properties.tabColor = GOLD

    ws.column_dimensions['A'].width = 4
    ws.column_dimensions['B'].width = 80

    r = 2
    ws.cell(row=r, column=2, value='SAFE HARBOR\u2122 EVIDENCE LEDGER').font = Font(name='Arial', size=16, bold=True, color=NAVY)
    r += 1
    ws.cell(row=r, column=2, value='Instructions & Quick Reference Guide').font = Font(name='Arial', size=12, color=GRAY_500)
    r += 1
    for c in range(1, 12):
        ws.cell(row=r, column=c).fill = PatternFill('solid', fgColor=GOLD)
    ws.row_dimensions[r].height = 3
    r += 2

    instructions = [
        ('WHAT IS THIS DOCUMENT?', 'This Evidence Ledger is the operational backbone of your Safe Harbor\u2122 compliance program. It documents every vendor, network endpoint, regulatory interaction, and quarterly audit \u2014 creating an irrefutable paper trail that proves "Reasonable Care" under Texas SB 1188.'),
        ('WHY IT MATTERS', 'During a state inquiry or Cure Notice, this ledger is what you submit to demonstrate active compliance governance. Without it, your Data Sovereignty Policy is a signed piece of paper. With it, you have a living, auditable record of due diligence that can stop a $250,000 fine or a $2,500/day Cure Notice penalty.'),
        ('THE GOLDEN RULE', '"If it isn\'t in the ledger, it didn\'t happen." Every vendor verification, every audit result, every regulatory interaction MUST be logged here. Verbal confirmations are worthless without written documentation.'),
        ('', ''),
        ('TAB 1: DIGITAL SUPPLY CHAIN', 'Document every vendor that touches patient data. Include their server location, verification proof, and sovereignty status. Send the Vendor Certification email template (from your AI Disclosure Kit) to every CRITICAL and HIGH tier vendor. Save their reply as a PDF and note it here.'),
        ('TAB 2: TECHNICAL RESIDENCY SIGNALS', 'Track your network infrastructure \u2014 routers, VPNs, remote access points, CDN endpoints. This tab proves that no data is being routed through foreign proxies or offshore IT support. Your IT provider should assist with this tab.'),
        ('TAB 3: REGULATORY & CURE NOTICE LOG', 'Log ALL regulatory contact, patient complaints about data privacy, and any Cure Notices. Document your response, attach evidence, and track resolution. This is your "Active Defense" register.'),
        ('TAB 4: QUARTERLY AUDIT TRAIL', 'Complete one row per quarter confirming that your Sentry Scan is clean, vendor certs are current, employee acknowledgments are signed, and your website disclosure is live. This rolling record is the strongest element of Safe Harbor defense.'),
        ('', ''),
        ('THE QUARTERLY SWEEP (Every 90 Days)', '1. Run a Sentry Watch scan at kairologic.com\n2. Review all vendor entries \u2014 update any that have changed\n3. Check for new employees needing acknowledgment forms\n4. Verify website AI disclosure is still live\n5. Log the audit in Tab 4\n6. File the updated ledger in your compliance folder'),
        ('BURDEN OF PROOF', 'In a Cure Notice scenario, the state gives you a window (typically 30 days) to prove compliance. This ledger, combined with your signed Data Sovereignty Policy, vendor certificates, and Sentry Scan reports, forms a complete evidentiary portfolio. Practices with this documentation have the strongest possible defense.'),
    ]

    for title, body in instructions:
        if title:
            ws.cell(row=r, column=2, value=title).font = Font(name='Arial', size=11, bold=True, color=NAVY)
            r += 1
        if body:
            ws.cell(row=r, column=2, value=body).font = Font(name='Arial', size=10, color=GRAY_700)
            ws.cell(row=r, column=2).alignment = Alignment(wrap_text=True, vertical='top')
            ws.row_dimensions[r].height = max(30, len(body) // 3)
            r += 1
        r += 1


def main():
    wb = Workbook()

    build_tab1_vendor_inventory(wb)
    build_tab2_residency_signals(wb)
    build_tab3_regulatory_log(wb)
    build_tab4_quarterly_audit(wb)
    build_tab5_instructions(wb)

    
    # Accept --output CLI arg
    output = None
    for i, arg in enumerate(sys.argv):
        if arg == '--output' and i + 1 < len(sys.argv):
            output = sys.argv[i + 1]
            import os
            os.makedirs(os.path.dirname(output) or '.', exist_ok=True)
    if not output:
        output = '/mnt/user-data/outputs/Evidence_Ledger.xlsx'

    wb.save(output)

    # Recalculate formulas
    try:
        result = subprocess.run(['python3', '/mnt/skills/public/xlsx/scripts/recalc.py', output, '30'],
                                capture_output=True, text=True, timeout=60)
        print(result.stdout)
    except Exception as e:
        print(f'Recalc note: {e}')

    import os
    print(f'XLSX generated: {output}')
    print(f'File size: {os.path.getsize(output) / 1024:.0f} KB')


if __name__ == '__main__':
    main()
